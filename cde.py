#Tworzenie dummy exel
from tracemalloc import start
from openpyxl import Workbook
from openpyxl import load_workbook

import datetime
import random

start = datetime.date(2020, 1, 1)
end = datetime.date(2020, 12, 31)

def random_date(start, end):
    """Generate a random datetime between `start` and `end`"""
    return start + datetime.timedelta(
        # Get a random amount of seconds between `start` and `end`
        seconds=random.randint(0, int((end - start).total_seconds())),
    )

wb = load_workbook("do-testu.xlsx")
ws = wb.active

# c = ws['A1':'A50']
for x in ws['A1':'A50']:
    for o in x:
        o.value = random_date(start, end)
        # print(o.value)

wb.save('do-testu.xlsx')
    

# print(random_date(start, end))


# #Lapanie calej kolumny
# column_a = ws['A']

# #Lapanie calej linni
# linni_1 = ws['1']

# #print wartosci z komurki
# for cell in column_a:
#     print(cell.value)


 