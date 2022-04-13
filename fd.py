#Filtr data
from openpyxl import load_workbook


wb = load_workbook("do-testu.xlsx")
ws = wb.active
<<<<<<< Updated upstream
lista_dat = []
for cell in ws:
    for o in cell:
        print(type(o.value))
        x = o.value.isoformat()
        lista_dat.append(x)

print(lista_dat)
=======
#wb.iso_dates = True to będzie chyba lepszy format daty

sheet_range  = ws["A1" : "A10"]


print(sheet_range.value)
>>>>>>> Stashed changes


        #Chce filtrowac wszystkie daty z tego pliku i 
        #sprawdziec z obecna data i wyciagnac cale komurki 
        #nastepnie zapisac do nowego pliku pod nazwa Wymaga konserwacji najlepiej ze zmienna dla daty.

